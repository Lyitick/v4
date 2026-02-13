"""Excel export REST API endpoint."""
from __future__ import annotations

import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

from Bot.database.get_db import get_db

from webapp.backend.dependencies import get_current_user

router = APIRouter()


def _build_excel(user_id: int, year: int, month: int) -> io.BytesIO:
    """Generate an .xlsx workbook with financial data."""
    # Use openpyxl; fall back to csv-in-xlsx if not installed
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    db = get_db()
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    def style_header(ws, cols: int) -> None:
        for col in range(1, cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Savings ──────────────────────────────
    ws_savings = wb.active
    ws_savings.title = "Накопления"
    ws_savings.append(["Категория", "Текущие", "Цель", "Назначение", "Прогресс %"])
    savings = db.get_all_savings_list(user_id)
    for s in savings:
        current = float(s.get("current", 0))
        goal = float(s.get("goal", 0))
        pct = round(current / goal * 100, 1) if goal > 0 else 0
        ws_savings.append([s["category"], current, goal, s.get("purpose", ""), pct])
    style_header(ws_savings, 5)
    ws_savings.column_dimensions["A"].width = 20
    ws_savings.column_dimensions["D"].width = 25

    # ── Sheet 2: Monthly Report ───────────────────────
    report = db.get_monthly_report_data(user_id, year, month)
    ws_report = wb.create_sheet("Отчёт за месяц")
    ws_report.append(["Показатель", "Сумма"])
    ws_report.append(["Месяц", report["month"]])
    ws_report.append(["Общий доход", report["total_income"]])
    ws_report.append(["Общий расход", report["total_expense"]])
    ws_report.append(["Баланс", round(report["total_income"] - report["total_expense"], 2)])
    ws_report.append(["Бытовые оплачено", report["household_paid"]])
    ws_report.append(["Бытовые всего", report["household_total"]])
    ws_report.append([])
    ws_report.append(["--- Доходы по категориям ---", ""])
    for item in report["income_by_category"]:
        ws_report.append([item["category"], item["amount"]])
    ws_report.append([])
    ws_report.append(["--- Расходы по категориям ---", ""])
    for item in report["expense_by_category"]:
        ws_report.append([item["category"], item["amount"]])
    style_header(ws_report, 2)
    ws_report.column_dimensions["A"].width = 30
    ws_report.column_dimensions["B"].width = 15

    # ── Sheet 3: Recurring Payments ───────────────────
    ws_recurring = wb.create_sheet("Повторяющиеся")
    ws_recurring.append(["Название", "Сумма", "Частота", "День месяца", "Следующая дата"])
    recurring = db.list_recurring_payments(user_id)
    for r in recurring:
        ws_recurring.append([r["title"], r["amount"], r["frequency"], r["day_of_month"], r.get("next_due_date", "")])
    style_header(ws_recurring, 5)
    ws_recurring.column_dimensions["A"].width = 25

    # ── Sheet 4: Household ────────────────────────────
    month_prefix = f"{year:04d}-{month:02d}"
    ws_household = wb.create_sheet("Бытовые платежи")
    ws_household.append(["Платёж", "Сумма", "Оплачено"])
    items = db.list_active_household_items(user_id)
    cursor = db.connection.cursor()
    for item in items:
        try:
            cursor.execute(
                f'SELECT is_paid FROM "{db.tables.household_payments}" WHERE user_id = ? AND month = ? AND question_code = ?',
                (user_id, month_prefix, item["code"]),
            )
            row = cursor.fetchone()
            is_paid = bool(row and int(row["is_paid"])) if row else False
        except Exception:
            is_paid = False
        ws_household.append([item["text"], item["amount"], "Да" if is_paid else "Нет"])
    style_header(ws_household, 3)
    ws_household.column_dimensions["A"].width = 35

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/excel")
async def export_excel(
    year: int = Query(default=None),
    month: int = Query(default=None),
    user: dict = Depends(get_current_user),
):
    """Export financial data as .xlsx file."""
    now = datetime.utcnow()
    year = year or now.year
    month = month or now.month

    buf = _build_excel(user["id"], year, month)
    filename = f"finance_report_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
